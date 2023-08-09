import openpyxl
import sys

# Check if the user provided a filename argument
if len(sys.argv) != 2:
    print("Usage: python script_name.py filename")
    sys.exit()

# Get the filename from the command-line argument
filename = sys.argv[1]

# Define the list of unacceptable curse words
curse_words = ['fuck', 'shit', 'bitch', 'cunt', 'cum', 'cock', 'pussy', 'god damnit', 'goddamnit', 'ass hole', 'asshole']

# Open the transcript file for reading
with open(filename, 'r') as transcript_file:
    transcript_lines = transcript_file.readlines()

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers to the worksheet
worksheet['A1'] = 'Timecode'
worksheet['B1'] = 'Speaker'
worksheet['C1'] = 'Line'

# Initialize the row counter
row_num = 2

# Search the transcript for each curse word and output identified sections to the worksheet
for i in range(0, len(transcript_lines), 4):
    timecode = transcript_lines[i].strip()
    speaker = transcript_lines[i+1].strip()
    line = transcript_lines[i+2].strip()
    
    # Use a flag variable to indicate if the curse word is found in the line
    curse_word_found = False

    for word in curse_words:
        if word.lower() in line.lower():
            curse_word_found = True
            break  # Exit the inner loop if a match is found

    if curse_word_found:
        worksheet.cell(row=row_num, column=1, value=timecode)
        worksheet.cell(row=row_num, column=2, value=speaker)
        worksheet.cell(row=row_num, column=3, value=line)
        row_num += 1

# Save the Excel workbook
workbook.save('identified_sections.xlsx')
